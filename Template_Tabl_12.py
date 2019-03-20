from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

wb = Workbook()

# Имя файла

File_Name = 'Tabl_2.xlsx'

# Активная вкладка в эксель таблице

ws = wb.active

# Стиль текста Название, Размер и Жирность

ft = Font(
            name='Courier New',
            size=9,
            bold=True
)

# Цыкл для всех ячеек таблицы задаёт размер и стиль текста

for x in range(1, 54):
    for y in range(1, 54):
        ws.cell(row=x, column=y).font = ft

# Стиль текста, Выравнивание горизонтально и вертикально по центру

at = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
)
# Цыкл для ячеек таблицы задаёт выравнивание по центру и перенос слов

for x in range(3, 54):
    for y in range(1, 54):
        ws.cell(row=x, column=y).alignment = at

at2 = Alignment(
            horizontal='center',
            vertical='center',
)

ws.cell(row=1, column=1).alignment = at2     # Стиль текста без переноса слов для верхних двух ячеек
ws.cell(row=2, column=12).alignment = at2

at3 = Alignment(                             # Стиль текста по правой стороне
            horizontal='right',
            vertical='center',
)

ws.cell(row=1, column=42).alignment = at3    # Ячейки с текстом по правой стороне
ws.cell(row=2, column=40).alignment = at3

tb = Border(                                 # Стиль Рамки для ячейки с каждой стороны
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
)

# width all column

ws.column_dimensions['A'].width = 2.7
ws.column_dimensions['B'].width = 2.7
ws.column_dimensions['C'].width = 2.7
ws.column_dimensions['D'].width = 3.6
ws.column_dimensions['E'].width = 4
ws.column_dimensions['F'].width = 2.7
ws.column_dimensions['G'].width = 2.7
ws.column_dimensions['H'].width = 2.7
ws.column_dimensions['I'].width = 2.7
ws.column_dimensions['J'].width = 2.7
ws.column_dimensions['K'].width = 2.7
ws.column_dimensions['L'].width = 2.7
ws.column_dimensions['M'].width = 2.7
ws.column_dimensions['N'].width = 2.7
ws.column_dimensions['O'].width = 2.7
ws.column_dimensions['P'].width = 2.7
ws.column_dimensions['Q'].width = 2.7
ws.column_dimensions['R'].width = 2.7
ws.column_dimensions['S'].width = 2.7
ws.column_dimensions['T'].width = 2.7
ws.column_dimensions['U'].width = 4
ws.column_dimensions['V'].width = 2.7
ws.column_dimensions['W'].width = 2.7
ws.column_dimensions['X'].width = 2.7
ws.column_dimensions['Y'].width = 2.7
ws.column_dimensions['Z'].width = 2.7

ws.column_dimensions['AA'].width = 2.7
ws.column_dimensions['AB'].width = 2.7
ws.column_dimensions['AC'].width = 2.7
ws.column_dimensions['AD'].width = 2.7
ws.column_dimensions['AE'].width = 2.7
ws.column_dimensions['AF'].width = 2.7
ws.column_dimensions['AG'].width = 2.7
ws.column_dimensions['AH'].width = 2.7
ws.column_dimensions['AI'].width = 2.7
ws.column_dimensions['AJ'].width = 2.7
ws.column_dimensions['AK'].width = 4
ws.column_dimensions['AL'].width = 2.7
ws.column_dimensions['AM'].width = 2.7
ws.column_dimensions['AN'].width = 2.7
ws.column_dimensions['AO'].width = 4
ws.column_dimensions['AP'].width = 2.7
ws.column_dimensions['AQ'].width = 2.7
ws.column_dimensions['AR'].width = 2.7
ws.column_dimensions['AS'].width = 4
ws.column_dimensions['AT'].width = 2.7
ws.column_dimensions['AU'].width = 4
ws.column_dimensions['AV'].width = 2.7
ws.column_dimensions['AW'].width = 4
ws.column_dimensions['AX'].width = 2.7
ws.column_dimensions['AY'].width = 2.7
ws.column_dimensions['AZ'].width = 2.3
ws.column_dimensions['BA'].width = 2.7

# Высота Рядков

ws.row_dimensions[1].height = 11.25
ws.row_dimensions[2].height = 11.25
ws.row_dimensions[3].height = 11.25
ws.row_dimensions[4].height = 11.25
ws.row_dimensions[5].height = 11.25
ws.row_dimensions[6].height = 11.25
ws.row_dimensions[7].height = 11.25
ws.row_dimensions[8].height = 11.25
ws.row_dimensions[9].height = 11.25
ws.row_dimensions[10].height = 11.25
ws.row_dimensions[11].height = 11.25
ws.row_dimensions[12].height = 11.25
ws.row_dimensions[13].height = 11.25
ws.row_dimensions[14].height = 11.25
ws.row_dimensions[15].height = 5.25
ws.row_dimensions[16].height = 11.25
ws.row_dimensions[17].height = 11.25
ws.row_dimensions[18].height = 11.25
ws.row_dimensions[19].height = 11.25
ws.row_dimensions[20].height = 11.25
ws.row_dimensions[21].height = 11.25
ws.row_dimensions[22].height = 11.25
ws.row_dimensions[23].height = 11.25
ws.row_dimensions[24].height = 11.25
ws.row_dimensions[25].height = 11.25
ws.row_dimensions[26].height = 5.25
ws.row_dimensions[27].height = 11.25
ws.row_dimensions[28].height = 11.25
ws.row_dimensions[29].height = 11.25
ws.row_dimensions[30].height = 11.25
ws.row_dimensions[31].height = 11.25
ws.row_dimensions[32].height = 11.25
ws.row_dimensions[33].height = 11.25
ws.row_dimensions[34].height = 11.25
ws.row_dimensions[35].height = 11.25
ws.row_dimensions[36].height = 11.25
ws.row_dimensions[37].height = 11.25
ws.row_dimensions[38].height = 3.75
ws.row_dimensions[39].height = 11.25
ws.row_dimensions[40].height = 11.25
ws.row_dimensions[41].height = 11.25
ws.row_dimensions[42].height = 3.75
ws.row_dimensions[43].height = 11.25
ws.row_dimensions[44].height = 11.25
ws.row_dimensions[45].height = 11.25
ws.row_dimensions[46].height = 11.25
ws.row_dimensions[47].height = 11.25
ws.row_dimensions[48].height = 11.25

# Обьединение ячеек

ws.merge_cells('A1:K1')
ws.merge_cells('AP1:AZ1')
ws.merge_cells('L2:AL2')
ws.merge_cells('AN2:AZ2')
ws.merge_cells('A3:D4')
ws.merge_cells('E3:AZ3')

ws.merge_cells('E4:H4')
ws.merge_cells('I4:L4')
ws.merge_cells('M4:P4')
ws.merge_cells('Q4:T4')
ws.merge_cells('U4:X4')
ws.merge_cells('Y4:AB4')
ws.merge_cells('AC4:AF4')
ws.merge_cells('AG4:AJ4')
ws.merge_cells('AK4:AN4')
ws.merge_cells('AO4:AR4')
ws.merge_cells('AS4:AV4')
ws.merge_cells('AW4:AZ4')

ws.merge_cells('A5:D5')
ws.merge_cells('E5:H5')
ws.merge_cells('I5:L5')
ws.merge_cells('M5:P5')
ws.merge_cells('Q5:T5')
ws.merge_cells('U5:X5')
ws.merge_cells('Y5:AB5')
ws.merge_cells('AC5:AF5')
ws.merge_cells('AG5:AJ5')
ws.merge_cells('AK5:AN5')
ws.merge_cells('AO5:AR5')
ws.merge_cells('AS5:AV5')
ws.merge_cells('AW5:AZ5')

ws.merge_cells('A6:D6')
ws.merge_cells('E6:H6')
ws.merge_cells('I6:L6')
ws.merge_cells('M6:P6')
ws.merge_cells('Q6:T6')
ws.merge_cells('U6:X6')
ws.merge_cells('Y6:AB6')
ws.merge_cells('AC6:AF6')
ws.merge_cells('AG6:AJ6')
ws.merge_cells('AK6:AN6')
ws.merge_cells('AO6:AR6')
ws.merge_cells('AS6:AV6')
ws.merge_cells('AW6:AZ6')

ws.merge_cells('A7:D7')
ws.merge_cells('E7:H7')
ws.merge_cells('I7:L7')
ws.merge_cells('M7:P7')
ws.merge_cells('Q7:T7')
ws.merge_cells('U7:X7')
ws.merge_cells('Y7:AB7')
ws.merge_cells('AC7:AF7')
ws.merge_cells('AG7:AJ7')
ws.merge_cells('AK7:AN7')
ws.merge_cells('AO7:AR7')
ws.merge_cells('AS7:AV7')
ws.merge_cells('AW7:AZ7')

ws.merge_cells('A8:D8')
ws.merge_cells('E8:H8')
ws.merge_cells('I8:L8')
ws.merge_cells('M8:P8')
ws.merge_cells('Q8:T8')
ws.merge_cells('U8:X8')
ws.merge_cells('Y8:AB8')
ws.merge_cells('AC8:AF8')
ws.merge_cells('AG8:AJ8')
ws.merge_cells('AK8:AN8')
ws.merge_cells('AO8:AR8')
ws.merge_cells('AS8:AV8')
ws.merge_cells('AW8:AZ8')

ws.merge_cells('A9:D9')
ws.merge_cells('E9:H9')
ws.merge_cells('I9:L9')
ws.merge_cells('M9:P9')
ws.merge_cells('Q9:T9')
ws.merge_cells('U9:X9')
ws.merge_cells('Y9:AB9')
ws.merge_cells('AC9:AF9')
ws.merge_cells('AG9:AJ9')
ws.merge_cells('AK9:AN9')
ws.merge_cells('AO9:AR9')
ws.merge_cells('AS9:AV9')
ws.merge_cells('AW9:AZ9')

ws.merge_cells('A10:D10')
ws.merge_cells('E10:H10')
ws.merge_cells('I10:L10')
ws.merge_cells('M10:P10')
ws.merge_cells('Q10:T10')
ws.merge_cells('U10:X10')
ws.merge_cells('Y10:AB10')
ws.merge_cells('AC10:AF10')
ws.merge_cells('AG10:AJ10')
ws.merge_cells('AK10:AN10')
ws.merge_cells('AO10:AR10')
ws.merge_cells('AS10:AV10')
ws.merge_cells('AW10:AZ10')

ws.merge_cells('A11:D11')
ws.merge_cells('E11:H11')
ws.merge_cells('I11:L11')
ws.merge_cells('M11:P11')
ws.merge_cells('Q11:T11')
ws.merge_cells('U11:X11')
ws.merge_cells('Y11:AB11')
ws.merge_cells('AC11:AF11')
ws.merge_cells('AG11:AJ11')
ws.merge_cells('AK11:AN11')
ws.merge_cells('AO11:AR11')
ws.merge_cells('AS11:AV11')
ws.merge_cells('AW11:AZ11')

ws.merge_cells('A12:D12')
ws.merge_cells('E12:H12')
ws.merge_cells('I12:L12')
ws.merge_cells('M12:P12')
ws.merge_cells('Q12:T12')
ws.merge_cells('U12:X12')
ws.merge_cells('Y12:AB12')
ws.merge_cells('AC12:AF12')
ws.merge_cells('AG12:AJ12')
ws.merge_cells('AK12:AN12')
ws.merge_cells('AO12:AR12')
ws.merge_cells('AS12:AV12')
ws.merge_cells('AW12:AZ12')

ws.merge_cells('A13:D13')
ws.merge_cells('E13:H13')
ws.merge_cells('I13:L13')
ws.merge_cells('M13:P13')
ws.merge_cells('Q13:T13')
ws.merge_cells('U13:X13')
ws.merge_cells('Y13:AB13')
ws.merge_cells('AC13:AF13')
ws.merge_cells('AG13:AJ13')
ws.merge_cells('AK13:AN13')
ws.merge_cells('AO13:AR13')
ws.merge_cells('AS13:AV13')
ws.merge_cells('AW13:AZ13')

ws.merge_cells('A14:D14')
ws.merge_cells('E14:H14')
ws.merge_cells('I14:L14')
ws.merge_cells('M14:P14')
ws.merge_cells('Q14:T14')
ws.merge_cells('U14:X14')
ws.merge_cells('Y14:AB14')
ws.merge_cells('AC14:AF14')
ws.merge_cells('AG14:AJ14')
ws.merge_cells('AK14:AN14')
ws.merge_cells('AO14:AR14')
ws.merge_cells('AS14:AV14')
ws.merge_cells('AW14:AZ14')

ws.merge_cells('A16:D16')
ws.merge_cells('E16:H16')
ws.merge_cells('I16:L16')
ws.merge_cells('M16:P16')
ws.merge_cells('Q16:T16')
ws.merge_cells('U16:X16')
ws.merge_cells('Y16:AB16')
ws.merge_cells('AC16:AF16')
ws.merge_cells('AG16:AJ16')
ws.merge_cells('AK16:AN16')
ws.merge_cells('AO16:AR16')
ws.merge_cells('AS16:AV16')
ws.merge_cells('AW16:AZ16')

ws.merge_cells('A17:D17')
ws.merge_cells('E17:H17')
ws.merge_cells('I17:L17')
ws.merge_cells('M17:P17')
ws.merge_cells('Q17:T17')
ws.merge_cells('U17:X17')
ws.merge_cells('Y17:AB17')
ws.merge_cells('AC17:AF17')
ws.merge_cells('AG17:AJ17')
ws.merge_cells('AK17:AN17')
ws.merge_cells('AO17:AR17')
ws.merge_cells('AS17:AV17')
ws.merge_cells('AW17:AZ17')

ws.merge_cells('A18:D18')
ws.merge_cells('E18:H18')
ws.merge_cells('I18:L18')
ws.merge_cells('M18:P18')
ws.merge_cells('Q18:T18')
ws.merge_cells('U18:X18')
ws.merge_cells('Y18:AB18')
ws.merge_cells('AC18:AF18')
ws.merge_cells('AG18:AJ18')
ws.merge_cells('AK18:AN18')
ws.merge_cells('AO18:AR18')
ws.merge_cells('AS18:AV18')
ws.merge_cells('AW18:AZ18')

ws.merge_cells('A19:D19')
ws.merge_cells('E19:H19')
ws.merge_cells('I19:L19')
ws.merge_cells('M19:P19')
ws.merge_cells('Q19:T19')
ws.merge_cells('U19:X19')
ws.merge_cells('Y19:AB19')
ws.merge_cells('AC19:AF19')
ws.merge_cells('AG19:AJ19')
ws.merge_cells('AK19:AN19')
ws.merge_cells('AO19:AR19')
ws.merge_cells('AS19:AV19')
ws.merge_cells('AW19:AZ19')

ws.merge_cells('A20:D20')
ws.merge_cells('E20:H20')
ws.merge_cells('I20:L20')
ws.merge_cells('M20:P20')
ws.merge_cells('Q20:T20')
ws.merge_cells('U20:X20')
ws.merge_cells('Y20:AB20')
ws.merge_cells('AC20:AF20')
ws.merge_cells('AG20:AJ20')
ws.merge_cells('AK20:AN20')
ws.merge_cells('AO20:AR20')
ws.merge_cells('AS20:AV20')
ws.merge_cells('AW20:AZ20')

ws.merge_cells('A21:D21')
ws.merge_cells('E21:H21')
ws.merge_cells('I21:L21')
ws.merge_cells('M21:P21')
ws.merge_cells('Q21:T21')
ws.merge_cells('U21:X21')
ws.merge_cells('Y21:AB21')
ws.merge_cells('AC21:AF21')
ws.merge_cells('AG21:AJ21')
ws.merge_cells('AK21:AN21')
ws.merge_cells('AO21:AR21')
ws.merge_cells('AS21:AV21')
ws.merge_cells('AW21:AZ21')

ws.merge_cells('A22:D22')
ws.merge_cells('E22:H22')
ws.merge_cells('I22:L22')
ws.merge_cells('M22:P22')
ws.merge_cells('Q22:T22')
ws.merge_cells('U22:X22')
ws.merge_cells('Y22:AB22')
ws.merge_cells('AC22:AF22')
ws.merge_cells('AG22:AJ22')
ws.merge_cells('AK22:AN22')
ws.merge_cells('AO22:AR22')
ws.merge_cells('AS22:AV22')
ws.merge_cells('AW22:AZ22')

ws.merge_cells('A23:D23')
ws.merge_cells('E23:H23')
ws.merge_cells('I23:L23')
ws.merge_cells('M23:P23')
ws.merge_cells('Q23:T23')
ws.merge_cells('U23:X23')
ws.merge_cells('Y23:AB23')
ws.merge_cells('AC23:AF23')
ws.merge_cells('AG23:AJ23')
ws.merge_cells('AK23:AN23')
ws.merge_cells('AO23:AR23')
ws.merge_cells('AS23:AV23')
ws.merge_cells('AW23:AZ23')

ws.merge_cells('A24:D24')
ws.merge_cells('E24:H24')
ws.merge_cells('I24:L24')
ws.merge_cells('M24:P24')
ws.merge_cells('Q24:T24')
ws.merge_cells('U24:X24')
ws.merge_cells('Y24:AB24')
ws.merge_cells('AC24:AF24')
ws.merge_cells('AG24:AJ24')
ws.merge_cells('AK24:AN24')
ws.merge_cells('AO24:AR24')
ws.merge_cells('AS24:AV24')
ws.merge_cells('AW24:AZ24')

ws.merge_cells('A25:D25')
ws.merge_cells('E25:H25')
ws.merge_cells('I25:L25')
ws.merge_cells('M25:P25')
ws.merge_cells('Q25:T25')
ws.merge_cells('U25:X25')
ws.merge_cells('Y25:AB25')
ws.merge_cells('AC25:AF25')
ws.merge_cells('AG25:AJ25')
ws.merge_cells('AK25:AN25')
ws.merge_cells('AO25:AR25')
ws.merge_cells('AS25:AV25')
ws.merge_cells('AW25:AZ25')

ws.merge_cells('A27:D27')
ws.merge_cells('E27:H27')
ws.merge_cells('I27:L27')
ws.merge_cells('M27:P27')
ws.merge_cells('Q27:T27')
ws.merge_cells('U27:X27')
ws.merge_cells('Y27:AB27')
ws.merge_cells('AC27:AF27')
ws.merge_cells('AG27:AJ27')
ws.merge_cells('AK27:AN27')
ws.merge_cells('AO27:AR27')
ws.merge_cells('AS27:AV27')
ws.merge_cells('AW27:AZ27')

ws.merge_cells('A28:D28')
ws.merge_cells('E28:H28')
ws.merge_cells('I28:L28')
ws.merge_cells('M28:P28')
ws.merge_cells('Q28:T28')
ws.merge_cells('U28:X28')
ws.merge_cells('Y28:AB28')
ws.merge_cells('AC28:AF28')
ws.merge_cells('AG28:AJ28')
ws.merge_cells('AK28:AN28')
ws.merge_cells('AO28:AR28')
ws.merge_cells('AS28:AV28')
ws.merge_cells('AW28:AZ28')

ws.merge_cells('A29:D29')
ws.merge_cells('E29:H29')
ws.merge_cells('I29:L29')
ws.merge_cells('M29:P29')
ws.merge_cells('Q29:T29')
ws.merge_cells('U29:X29')
ws.merge_cells('Y29:AB29')
ws.merge_cells('AC29:AF29')
ws.merge_cells('AG29:AJ29')
ws.merge_cells('AK29:AN29')
ws.merge_cells('AO29:AR29')
ws.merge_cells('AS29:AV29')
ws.merge_cells('AW29:AZ29')

ws.merge_cells('A30:D30')
ws.merge_cells('E30:H30')
ws.merge_cells('I30:L30')
ws.merge_cells('M30:P30')
ws.merge_cells('Q30:T30')
ws.merge_cells('U30:X30')
ws.merge_cells('Y30:AB30')
ws.merge_cells('AC30:AF30')
ws.merge_cells('AG30:AJ30')
ws.merge_cells('AK30:AN30')
ws.merge_cells('AO30:AR30')
ws.merge_cells('AS30:AV30')
ws.merge_cells('AW30:AZ30')

ws.merge_cells('A31:D31')
ws.merge_cells('E31:H31')
ws.merge_cells('I31:L31')
ws.merge_cells('M31:P31')
ws.merge_cells('Q31:T31')
ws.merge_cells('U31:X31')
ws.merge_cells('Y31:AB31')
ws.merge_cells('AC31:AF31')
ws.merge_cells('AG31:AJ31')
ws.merge_cells('AK31:AN31')
ws.merge_cells('AO31:AR31')
ws.merge_cells('AS31:AV31')
ws.merge_cells('AW31:AZ31')

ws.merge_cells('A32:D32')
ws.merge_cells('E32:H32')
ws.merge_cells('I32:L32')
ws.merge_cells('M32:P32')
ws.merge_cells('Q32:T32')
ws.merge_cells('U32:X32')
ws.merge_cells('Y32:AB32')
ws.merge_cells('AC32:AF32')
ws.merge_cells('AG32:AJ32')
ws.merge_cells('AK32:AN32')
ws.merge_cells('AO32:AR32')
ws.merge_cells('AS32:AV32')
ws.merge_cells('AW32:AZ32')

ws.merge_cells('A33:D33')
ws.merge_cells('E33:H33')
ws.merge_cells('I33:L33')
ws.merge_cells('M33:P33')
ws.merge_cells('Q33:T33')
ws.merge_cells('U33:X33')
ws.merge_cells('Y33:AB33')
ws.merge_cells('AC33:AF33')
ws.merge_cells('AG33:AJ33')
ws.merge_cells('AK33:AN33')
ws.merge_cells('AO33:AR33')
ws.merge_cells('AS33:AV33')
ws.merge_cells('AW33:AZ33')

ws.merge_cells('A34:D34')
ws.merge_cells('E34:H34')
ws.merge_cells('I34:L34')
ws.merge_cells('M34:P34')
ws.merge_cells('Q34:T34')
ws.merge_cells('U34:X34')
ws.merge_cells('Y34:AB34')
ws.merge_cells('AC34:AF34')
ws.merge_cells('AG34:AJ34')
ws.merge_cells('AK34:AN34')
ws.merge_cells('AO34:AR34')
ws.merge_cells('AS34:AV34')
ws.merge_cells('AW34:AZ34')

ws.merge_cells('A35:D35')
ws.merge_cells('E35:H35')
ws.merge_cells('I35:L35')
ws.merge_cells('M35:P35')
ws.merge_cells('Q35:T35')
ws.merge_cells('U35:X35')
ws.merge_cells('Y35:AB35')
ws.merge_cells('AC35:AF35')
ws.merge_cells('AG35:AJ35')
ws.merge_cells('AK35:AN35')
ws.merge_cells('AO35:AR35')
ws.merge_cells('AS35:AV35')
ws.merge_cells('AW35:AZ35')

ws.merge_cells('A36:D36')
ws.merge_cells('E36:H36')
ws.merge_cells('I36:L36')
ws.merge_cells('M36:P36')
ws.merge_cells('Q36:T36')
ws.merge_cells('U36:X36')
ws.merge_cells('Y36:AB36')
ws.merge_cells('AC36:AF36')
ws.merge_cells('AG36:AJ36')
ws.merge_cells('AK36:AN36')
ws.merge_cells('AO36:AR36')
ws.merge_cells('AS36:AV36')
ws.merge_cells('AW36:AZ36')

ws.merge_cells('A37:D37')
ws.merge_cells('E37:H37')
ws.merge_cells('I37:L37')
ws.merge_cells('M37:P37')
ws.merge_cells('Q37:T37')
ws.merge_cells('U37:X37')
ws.merge_cells('Y37:AB37')
ws.merge_cells('AC37:AF37')
ws.merge_cells('AG37:AJ37')
ws.merge_cells('AK37:AN37')
ws.merge_cells('AO37:AR37')
ws.merge_cells('AS37:AV37')
ws.merge_cells('AW37:AZ37')

ws.merge_cells('A39:D39')
ws.merge_cells('E39:H39')
ws.merge_cells('I39:L39')
ws.merge_cells('M39:P39')
ws.merge_cells('Q39:T39')
ws.merge_cells('U39:X39')
ws.merge_cells('Y39:AB39')
ws.merge_cells('AC39:AF39')
ws.merge_cells('AG39:AJ39')
ws.merge_cells('AK39:AN39')
ws.merge_cells('AO39:AR39')
ws.merge_cells('AS39:AV39')
ws.merge_cells('AW39:AZ39')

ws.merge_cells('A40:D40')
ws.merge_cells('E40:H40')
ws.merge_cells('I40:L40')
ws.merge_cells('M40:P40')
ws.merge_cells('Q40:T40')
ws.merge_cells('U40:X40')
ws.merge_cells('Y40:AB40')
ws.merge_cells('AC40:AF40')
ws.merge_cells('AG40:AJ40')
ws.merge_cells('AK40:AN40')
ws.merge_cells('AO40:AR40')
ws.merge_cells('AS40:AV40')
ws.merge_cells('AW40:AZ40')

ws.merge_cells('A41:D41')
ws.merge_cells('E41:H41')
ws.merge_cells('I41:L41')
ws.merge_cells('M41:P41')
ws.merge_cells('Q41:T41')
ws.merge_cells('U41:X41')
ws.merge_cells('Y41:AB41')
ws.merge_cells('AC41:AF41')
ws.merge_cells('AG41:AJ41')
ws.merge_cells('AK41:AN41')
ws.merge_cells('AO41:AR41')
ws.merge_cells('AS41:AV41')
ws.merge_cells('AW41:AZ41')

#  нижняя панель (подвал)

ws.merge_cells('A43:D45')
ws.merge_cells('E43:H45')

ws.merge_cells('I43:W43')
ws.merge_cells('I44:K45')
ws.merge_cells('L44:S44')
ws.merge_cells('L45:O45')
ws.merge_cells('P45:S45')
ws.merge_cells('T44:W45')

ws.merge_cells('X43:AL43')
ws.merge_cells('X44:Z45')
ws.merge_cells('AA44:AH44')
ws.merge_cells('AA45:AD45')
ws.merge_cells('AE45:AH45')
ws.merge_cells('AI44:AL45')

ws.merge_cells('AM43:BA43')
ws.merge_cells('AM44:AO45')
ws.merge_cells('AP44:AW44')
ws.merge_cells('AP45:AS45')
ws.merge_cells('AT45:AW45')
ws.merge_cells('AX44:BA45')

ws.merge_cells('A46:D46')
ws.merge_cells('E46:H46')
ws.merge_cells('I46:K46')
ws.merge_cells('L46:O46')
ws.merge_cells('P46:S46')
ws.merge_cells('T46:W46')
ws.merge_cells('X46:Z46')
ws.merge_cells('AA46:AD46')
ws.merge_cells('AE46:AH46')
ws.merge_cells('AI46:AL46')
ws.merge_cells('AM46:AO46')
ws.merge_cells('AP46:AS46')
ws.merge_cells('AT46:AW46')
ws.merge_cells('AX46:BA46')

ws.merge_cells('A47:D52')
ws.merge_cells('E47:H47')
ws.merge_cells('I47:K47')
ws.merge_cells('L47:O47')
ws.merge_cells('P47:S47')
ws.merge_cells('T47:W47')
ws.merge_cells('X47:Z47')
ws.merge_cells('AA47:AD47')
ws.merge_cells('AE47:AH47')
ws.merge_cells('AI47:AL47')
ws.merge_cells('AM47:AO47')
ws.merge_cells('AP47:AS47')
ws.merge_cells('AT47:AW47')
ws.merge_cells('AX47:BA47')

for x in range(3, 5):     # Рамка верхняя, вокруг месяцев
    for y in range(1, 53):
        ws.cell(row=x, column=y).border = tb

for x in range(43, 46):   # Рамка нижняя вокруг годовых данных
    for y in range(1, 54):
        ws.cell(row=x, column=y).border = tb

# Базовый текст таблицы

ws['A1'] = 'Таблиця 1.2. Рівень води, см'

ws['A3'] = 'Число'
ws['E3'] = 'Місяць'
ws['E4'] = '1'
ws['I4'] = '2'
ws['M4'] = '3'
ws['Q4'] = '4'
ws['U4'] = '5'
ws['Y4'] = '6'
ws['AC4'] = '7'
ws['AG4'] = '8'
ws['AK4'] = '9'
ws['AO4'] = '10'
ws['AS4'] = '11'
ws['AW4'] = '12'

ws['A5'] = '1'
ws['A6'] = '2'
ws['A7'] = '3'
ws['A8'] = '4'
ws['A9'] = '5'
ws['A10'] = '6'
ws['A11'] = '7'
ws['A12'] = '8'
ws['A13'] = '9'
ws['A14'] = '10'

ws['A16'] = '11'
ws['A17'] = '12'
ws['A18'] = '13'
ws['A19'] = '14'
ws['A20'] = '15'
ws['A21'] = '16'
ws['A22'] = '17'
ws['A23'] = '18'
ws['A24'] = '19'
ws['A25'] = '20'

ws['A27'] = '21'
ws['A28'] = '22'
ws['A29'] = '23'
ws['A30'] = '24'
ws['A31'] = '25'
ws['A32'] = '26'
ws['A33'] = '27'
ws['A34'] = '28'
ws['A35'] = '29'
ws['A36'] = '30'
ws['A37'] = '31'

ws['A39'] = 'Середній'
ws['A40'] = 'Вищий'
ws['A41'] = 'Нижчий'

ws['E43'] = 'Середній рівень'

ws['I43'] = 'Вищий'
ws['I44'] = 'рівень'
ws['L44'] = 'дата'
ws['L45'] = 'перша'
ws['P45'] = 'остання'
ws['T44'] = 'число випадків'

ws['X43'] = 'Нижчий періоду відкритого русла'
ws['X44'] = 'рівень'
ws['AA44'] = 'дата'
ws['AA45'] = 'перша'
ws['AE45'] = 'остання'
ws['AI44'] = 'число випадків'

ws['AM43'] = 'Нижчий зимового періоду'
ws['AM44'] = 'рівень'
ws['AP44'] = 'дата'
ws['AP45'] = 'перша'
ws['AT45'] = 'остання'
ws['AX44'] = 'число випадків'

ws['A46'] = 'За рік'

wb.save(filename = File_Name)